from collections import defaultdict
from datetime import datetime
from io import BytesIO
from logger import setup_logger
from database.db_helper import get_db
from database.models import (
    SurveyAnswer,
    Broadcast,
    SurveyResponse,
    Chat,
    ChatMember,
)
from database.chat_service import (
    get_active_memberships_with_titles,
    get_all_active_chat_members,
)
from database.user_service import get_all_users
from database.task_service import get_all_tasks
from database.homework_service import get_all_homeworks
from timezone_utils import format_moscow
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

logger = setup_logger(__name__)


def _write_headers(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)


_ROLE_LABELS = {"mentor": "Наставник", "student": "Участник"}


def _build_users_sheet(wb: Workbook, ws, users: list, tasks: list, homeworks: list) -> None:
    user_tg_ids: set[int] = {u.tg_id for u in users if u.tg_id is not None}

    tasks_by_student: dict[int, list] = defaultdict(list)
    for t in tasks:
        tasks_by_student[t.student_id].append(t)

    hws_by_student: dict[int, list] = defaultdict(list)
    for h in homeworks:
        hws_by_student[h.student_id].append(h)

    memberships = get_active_memberships_with_titles()
    chats_by_tg_id: dict[int, list[str]] = defaultdict(list)
    for tg_id, chat_title in memberships:
        if chat_title:
            chats_by_tg_id[tg_id].append(chat_title)

    all_members = get_all_active_chat_members()
    seen_tg_ids: dict[int, object] = {}
    for m in all_members:
        if m.user_tg_id not in seen_tg_ids:
            seen_tg_ids[m.user_tg_id] = m
    unregistered = [
        m for tg_id, m in seen_tg_ids.items() if tg_id not in user_tg_ids
    ]

    headers = [
        "ID",
        "Telegram ID",
        "TG ник",
        "Роль",
        "Зарегистрирован в боте",
        "Имя",
        "Фамилия",
        "Дата создания",
        "Дата регистрации",
        "Всего задач",
        "Всего ДЗ",
        "Чаты",
    ]
    _write_headers(ws, headers)

    row_num = 2

    for user in users:
        user_tasks = tasks_by_student.get(user.id, [])
        user_hws = hws_by_student.get(user.id, [])
        chats_str = ", ".join(chats_by_tg_id.get(user.tg_id, [])) if user.tg_id else ""

        row = [
            user.id,
            user.tg_id or "",
            f"@{user.tg_nickname}" if user.tg_nickname else "",
            _ROLE_LABELS.get(user.role.value, user.role.value),
            "Да",
            user.first_name or "",
            user.last_name or "",
            format_moscow(user.created_at, "%d.%m.%Y %H:%M") if user.created_at else "",
            format_moscow(user.registered_at, "%d.%m.%Y %H:%M") if user.registered_at else "",
            len(user_tasks),
            len(user_hws),
            chats_str,
        ]
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_idx, value=value)
        row_num += 1

    for member in unregistered:
        chats_str = ", ".join(chats_by_tg_id.get(member.user_tg_id, []))
        row = [
            "",
            member.user_tg_id,
            f"@{member.username}" if member.username else "",
            "",
            "Нет",
            member.first_name or "",
            member.last_name or "",
            "",
            "",
            "",
            "",
            chats_str,
        ]
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_idx, value=value)
        row_num += 1

    _auto_width(ws)
    logger.info(f"Users sheet: {row_num - 2} rows ({len(users)} registered, {len(unregistered)} unregistered)")


def _build_survey_sheet(wb: Workbook, users: list) -> None:
    ws = wb.create_sheet("Опросы")

    user_tg_ids: set[int] = {u.tg_id for u in users if u.tg_id is not None}
    user_by_tg_id = {u.tg_id: u for u in users if u.tg_id is not None}

    question_name_map = {
        "q1": "Ответ 1",
        "q1_followup_low": "Фолоу-ап к ответу 1",
        "q1_followup_mid": "Фолоу-ап к ответу 1",
        "q1_followup_high": "Фолоу-ап к ответу 1",
        "q2": "Ответ 2",
        "q2_followup_low": "Фолоу-ап к ответу 2",
        "q2_followup_mid": "Фолоу-ап к ответу 2",
        "q2_followup_high": "Фолоу-ап к ответу 2",
        "q3": "Ответ 3",
        "q3_followup_low": "Фолоу-ап к ответу 3",
        "q3_followup_mid": "Фолоу-ап к ответу 3",
        "q3_followup_high": "Фолоу-ап к ответу 3",
        "q4": "Ответ 4",
        "q4_followup_low": "Фолоу-ап к ответу 4",
        "q4_followup_mid": "Фолоу-ап к ответу 4",
        "q4_followup_high": "Фолоу-ап к ответу 4",
    }

    with get_db() as db:
        question_keys = [
            q[0] for q in (
                db.query(SurveyAnswer.question_key)
                .distinct()
                .order_by(SurveyAnswer.question_key)
                .all()
            )
        ]

        broadcasts = (
            db.query(Broadcast)
            .join(SurveyResponse, Broadcast.id == SurveyResponse.broadcast_id)
            .distinct()
            .all()
        )
        broadcast_ids = [b.id for b in broadcasts]

        responses = (
            db.query(SurveyResponse)
            .filter(SurveyResponse.broadcast_id.in_(broadcast_ids))
            .all()
        ) if broadcast_ids else []

        response_ids = [r.id for r in responses]
        answers_all = (
            db.query(SurveyAnswer)
            .filter(SurveyAnswer.response_id.in_(response_ids))
            .all()
        ) if response_ids else []

        chat_ids = {r.chat_id for r in responses}
        chats = (
            db.query(Chat)
            .filter(Chat.id.in_(chat_ids))
            .all()
        ) if chat_ids else []
        members = (
            db.query(ChatMember)
            .filter(ChatMember.chat_id.in_(chat_ids))
            .all()
        ) if chat_ids else []

    headers = [
        "Айди опроса",
        "Дата отправки",
        "TG ник куратора",
        "Название чата",
        "TG ID пользователя",
        "TG ник пользователя",
        "Зарегистрирован в боте",
        "Имя",
        "Фамилия",
        "Дата окончания опроса",
        "Статус",
    ]
    for q_key in question_keys:
        headers.append(question_name_map.get(q_key, q_key))
    _write_headers(ws, headers)

    responses_by_broadcast: dict[int, list] = defaultdict(list)
    for r in responses:
        responses_by_broadcast[r.broadcast_id].append(r)

    answers_by_response: dict[int, dict] = defaultdict(dict)
    for a in answers_all:
        answers_by_response[a.response_id][a.question_key] = a

    chat_by_id = {c.id: c for c in chats}
    member_by_chat_user: dict[tuple, object] = {}
    for m in members:
        member_by_chat_user[(m.chat_id, m.user_tg_id)] = m

    row_num = 2
    for broadcast in broadcasts:
        curator_user = user_by_tg_id.get(broadcast.curator_tg_id)
        curator_username = (
            f"@{curator_user.tg_nickname}"
            if curator_user and curator_user.tg_nickname
            else str(broadcast.curator_tg_id)
        )

        for response in responses_by_broadcast[broadcast.id]:
            chat = chat_by_id.get(response.chat_id)
            chat_title = chat.chat_title if chat else f"Chat {response.chat_id}"

            member = member_by_chat_user.get((response.chat_id, response.user_tg_id))
            username = member.username if member else None
            first_name = member.first_name if member else None
            last_name = member.last_name if member else None

            answer_dict = answers_by_response.get(response.id, {})

            row_data = [
                broadcast.id,
                format_moscow(broadcast.sent_at, "%d.%m.%Y %H:%M") if broadcast.sent_at else "",
                curator_username,
                chat_title,
                response.user_tg_id,
                username or "",
                "Да" if response.user_tg_id in user_tg_ids else "Нет",
                first_name or "",
                last_name or "",
                format_moscow(response.completed_at, "%d.%m.%Y %H:%M") if response.completed_at else "",
                "Завершен" if response.is_completed else "Не завершен",
            ]

            for q_key in question_keys:
                answer = answer_dict.get(q_key)
                if answer:
                    if answer.answer_value is not None:
                        row_data.append(answer.answer_value)
                    elif answer.answer_text:
                        row_data.append(answer.answer_text)
                    else:
                        row_data.append("")
                else:
                    row_data.append("")

            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_num, column=col_idx, value=value)
            row_num += 1

    _auto_width(ws)
    logger.info(f"Survey sheet: {row_num - 2} rows")


_TASK_STATUS_LABELS = {
    "unchecked": "Не проверено",
    "approved": "Одобрено",
    "disapproved": "Отклонено",
    "postponed": "Отложено",
}


def _build_tasks_sheet(wb: Workbook, users: list, tasks: list) -> None:
    ws = wb.create_sheet("Отбор")

    user_map = {u.id: u for u in users}

    headers = [
        "ID задачи",
        "ID лида",
        "Статус",
        "Участник",
        "TG ник участника",
        "Наставник",
        "Дата создания",
        "Дата обновления",
    ]
    _write_headers(ws, headers)

    for row_num, task in enumerate(tasks, start=2):
        student = user_map.get(task.student_id)
        mentor = user_map.get(task.mentor_id)

        student_name = " ".join(filter(None, [student.first_name, student.last_name])) if student else ""
        student_nick = f"@{student.tg_nickname}" if student and student.tg_nickname else ""
        mentor_nick = f"@{mentor.tg_nickname}" if mentor and mentor.tg_nickname else ""

        row = [
            task.id,
            task.lead_id,
            _TASK_STATUS_LABELS.get(task.status.value, task.status.value),
            student_name,
            student_nick,
            mentor_nick,
            format_moscow(task.created_at, "%d.%m.%Y %H:%M") if task.created_at else "",
            format_moscow(task.updated_at, "%d.%m.%Y %H:%M") if task.updated_at else "",
        ]
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_idx, value=value)

    _auto_width(ws)
    logger.info(f"Tasks sheet: {len(tasks)} rows")


_HW_STATUS_LABELS = {
    "pending": "Ожидает",
    "in_progress": "В процессе",
    "submitted": "Сдано",
    "pending_mentor": "На проверке",
    "postponed": "Отложено",
    "approved": "Одобрено",
    "edit": "На доработке",
    "edit_from_mentor": "На доработке (ментор)",
}


def _build_homework_sheet(wb: Workbook, users: list, homeworks: list) -> None:
    ws = wb.create_sheet("Домашние задания")

    user_map = {u.id: u for u in users}

    headers = [
        "ID ДЗ",
        "ID лида",
        "Статус",
        "Участник",
        "TG ник участника",
        "Наставник",
        "Вопрос 1",
        "Вопрос 2",
        "Вопрос 3",
        "Вопрос 4",
        "Вопрос 5",
        "Дедлайн",
        "Обратная связь",
        "Оценка",
        "Причина возврата",
        "Дата создания",
        "Дата обновления",
    ]
    _write_headers(ws, headers)

    for row_num, hw in enumerate(homeworks, start=2):
        student = user_map.get(hw.student_id)
        mentor = user_map.get(hw.mentor_id) if hw.mentor_id is not None else None

        student_name = " ".join(filter(None, [student.first_name, student.last_name])) if student else ""
        student_nick = f"@{student.tg_nickname}" if student and student.tg_nickname else ""
        mentor_nick = f"@{mentor.tg_nickname}" if mentor and mentor.tg_nickname else ""

        row = [
            hw.id,
            hw.lead_id,
            _HW_STATUS_LABELS.get(hw.status.value, hw.status.value),
            student_name,
            student_nick,
            mentor_nick,
            hw.first_hw or "",
            hw.second_hw or "",
            hw.third_hw or "",
            hw.fourth_hw or "",
            hw.fifth_hw or "",
            format_moscow(hw.deadline, "%d.%m.%Y %H:%M") if hw.deadline else "",
            hw.feedback or "",
            hw.rating if hw.rating is not None else "",
            hw.edit_reason_from_mentor or "",
            format_moscow(hw.created_at, "%d.%m.%Y %H:%M") if hw.created_at else "",
            format_moscow(hw.updated_at, "%d.%m.%Y %H:%M") if hw.updated_at else "",
        ]
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=col_idx, value=value)

    _auto_width(ws)
    logger.info(f"Homework sheet: {len(homeworks)} rows")


def generate_survey_export() -> BytesIO:
    """Generate XLSX export with all data sheets."""
    users = get_all_users()
    tasks = get_all_tasks()
    homeworks = get_all_homeworks()

    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"

    _build_users_sheet(wb, ws, users, tasks, homeworks)
    _build_survey_sheet(wb, users)
    _build_tasks_sheet(wb, users, tasks)
    _build_homework_sheet(wb, users, homeworks)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
