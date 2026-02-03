from io import BytesIO
from datetime import datetime
from logger import setup_logger
from database.db_helper import get_db
from database.models import SurveyAnswer
from database.broadcast_service import (
    get_all_broadcasts_with_responses,
    get_broadcast_responses,
    get_response_answers,
)
from database.chat_service import get_chat_by_db_id, get_chat_member_by_user_tg_id
from database.user_service import find_by_tg_id
from timezone_utils import format_moscow
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

logger = setup_logger(__name__)


def generate_survey_export() -> BytesIO:
    """Generate XLSX export with all broadcast data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Broadcast Data"

    # Headers
    headers = [
        "Айди опроса",
        "Дата отправки",
        "TG ник куратора",
        "Название чата",
        "TG ID пользователя",
        "TG ник пользователя",
        "Имя",
        "Фамилия",
        "Дата окончания опроса",
        "Статус",
    ]

    # Add question columns dynamically
    with get_db() as db:
        # Get all unique question keys
        question_keys = (
            db.query(SurveyAnswer.question_key)
            .distinct()
            .order_by(SurveyAnswer.question_key)
            .all()
        )
        question_keys = [q[0] for q in question_keys]

        # Map question keys to Russian column names
        question_name_map = {
            "q1": "Ответ 1",
            "q1_followup_low": "Фолоу-ап к ответу 1",
            "q1_followup_mid": "Фолоу-ап к ответу 1",
            "q1_followup_high": "Фолоу-ап к ответу 1",
            "q2": "Ответ 2",
            "q2_followup_low": "Фолоу-ап к ответу 2",
            "q2_followup_mid": "Фолоу-ап к ответу 2",
            "q2_followup_high": "Фолоу-ап к ответу 2",
            "q3": "Ответ 3",
            "q3_followup_low": "Фолоу-ап к ответу 3",
            "q3_followup_mid": "Фолоу-ап к ответу 3",
            "q3_followup_high": "Фолоу-ап к ответу 3",
            "q4": "Ответ 4",
            "q4_followup_low": "Фолоу-ап к ответу 4",
            "q4_followup_mid": "Фолоу-ап к ответу 4",
            "q4_followup_high": "Фолоу-ап к ответу 4",
        }

        for q_key in question_keys:
            # Use mapped name if available, otherwise use the key
            column_name = question_name_map.get(q_key, q_key)
            headers.append(column_name)

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Get all broadcasts with responses
    broadcasts = get_all_broadcasts_with_responses()

    row_num = 2
    for broadcast in broadcasts:
        # Get responses for this broadcast
        responses = get_broadcast_responses(broadcast.id)

        for response in responses:
            # Get chat
            chat = get_chat_by_db_id(response.chat_id)
            chat_title = chat.chat_title if chat else f"Chat {response.chat_id}"

            # Get user info from chat members
            username = None
            first_name = None
            last_name = None
            if chat:
                member = get_chat_member_by_user_tg_id(chat.id, response.user_tg_id)
                if member:
                    username = member.username
                    first_name = member.first_name
                    last_name = member.last_name

            # Get answers for this response
            answers = get_response_answers(response.id)
            answer_dict = {a.question_key: a for a in answers}

            # Get curator username from User table
            curator_user = find_by_tg_id(broadcast.curator_tg_id)
            curator_username = f"@{curator_user.tg_nickname}" if curator_user and curator_user.tg_nickname else str(broadcast.curator_tg_id)
            
            # Build row data
            row_data = [
                broadcast.id,
                format_moscow(broadcast.sent_at, "%d.%m.%Y %H:%M") if broadcast.sent_at else "",
                curator_username,
                chat_title,
                response.user_tg_id,
                username or "",
                first_name or "",
                last_name or "",
                format_moscow(response.completed_at, "%d.%m.%Y %H:%M") if response.completed_at else "",
                "Завершен" if response.is_completed else "Не завершен",
            ]

            # Add answers
            for q_key in question_keys:
                answer = answer_dict.get(q_key)
                if answer:
                    if answer.answer_value is not None:
                        row_data.append(answer.answer_value)
                    elif answer.answer_text:
                        row_data.append(answer.answer_text)
                    else:
                        row_data.append("")
                else:
                    row_data.append("")

            # Write row
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_num, column=col_idx, value=value)

            row_num += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info(f"Generated XLSX export with {row_num - 2} rows")
    return output
